"""Work with excel sheets
- using openpyxl: module utils for excel sheets


Note for openpyxl vocabulary:
- workbook represents the file excel sheet ( containing inner sheets )
- > worksheet is the sheet within the excel file sheet
"""
import re
from pathlib import Path

import openpyxl

# from custom_utils import log, log_header

CURRENT_DIR = Path(__file__).parent
LOCAL_EXCEL_FOLDER = CURRENT_DIR / 'samples'
LOCAL_EXCEL_NAME = LOCAL_EXCEL_FOLDER / 'local_excel_sheet.xlsx'
LOCAL_EXCEL_PATH = CURRENT_DIR / LOCAL_EXCEL_NAME

LOCAL_SAMPLE_EMPLOYEES_RANTING = LOCAL_EXCEL_FOLDER / 'Employee Ratings.xlsx'


def create_local_excel_file(filename=LOCAL_EXCEL_PATH, sheet_name='Sheet 1'):
    '''Create local exel file  - called workbook Then saves it'''
    print('\n[[ OpenPyXl ] Automate excel spreadsheet - Create/Load Spreadsheet ]')
    if not filename.exists():
        workbook = openpyxl.Workbook()
        workbook.save(filename)
    else:
        workbook = openpyxl.load_workbook(filename)

    workbook.create_sheet(sheet_name)
    worksheet = workbook[sheet_name]

    return workbook, worksheet, filename

# Reads excel content


def automate_excel_spreadsheet_import(filename=LOCAL_EXCEL_PATH, sheet_name="Sheet"):
    """Import content of the excel sheet"""
    print('\n[[ OpenPyXl ] Automate excel spreadsheet - Edit Spreadsheet Sheet Cell ]')

    # workbook: represents excel sheet with openpyxl
    workbook = openpyxl.load_workbook(filename)

    workbook.save(filename)

    # active first excel file sheet - where we add content
    current_worksheet = workbook[sheet_name]

    # cell targeting
    cell_a1 = current_worksheet['A1']

    # Modifying the cell
    current_worksheet['A1'] = 'Hello Excel'

    print('\t> [ ℹ️  workbook ]\t\t\t', workbook)
    print('\t> [ ℹ️  current worksheet ]\t\t', current_worksheet)
    print('\t> [ ℹ️  current_worksheet cell_a1 ]\t', cell_a1)
    print('\t> [ ℹ️  current_worksheet content ]\t',
          list(current_worksheet.values), '\n'
          )
    return workbook


def create_spreadsheet_sheet(workbook_path, new_sheet_name):
    '''create_spreadsheet_sheet Creates a spreadsheet's sheet

    Args:
                    workbook_path (str): path of the local file
                    new_sheet_name (str): name of the sheet to create

    Returns:
                    Worksheet: The spreadsheet's worksheet
    '''
    print('\n[[ OpenPyXl ] Automate excel spreadsheet - Create Spreadsheet Sheet ]')

    # Loads the spreadsheet file as a workbook
    workbook = openpyxl.load_workbook(workbook_path)

    # Create a spreadsheet's sheet ( after the first sheet )
    if any(ws.title != new_sheet_name for ws in workbook.worksheets):
        print(f'\t> [ Sheet Creation ] Creating new sheet "{new_sheet_name}"')
        workbook.create_sheet(new_sheet_name, 1)
    else:
        print(
            f'\t > [ Sheet creation ] Sheet "{
                new_sheet_name}" already exist'
        )

    # Save modification
    workbook.save(workbook_path)

    # Returns a sheet
    return workbook, workbook[new_sheet_name], workbook_path


def rename_spreadsheet_sheet(filepath, sheet_name, new_sheet_name):
    '''rename_spreadsheet_sheet Renames Spreadsheet sheet

    Args:
                    filepath (str): path of the excel spreadsheet file
                    sheet_name (str): name of the spreadsheet sheet
                    new_sheet_name (str): new name for the spreadsheet sheet

    Returns:
                    Workbook: the modified workbook
    '''
    print('\n[[ OpenPyXl ] Automate excel spreadsheet - Rename Spreadsheet sheet ]')
    workbook = openpyxl.load_workbook(filepath)
    try:
        if not workbook[sheet_name]:
            print(
                f'\t> [ Sheet Rename ] ❌ Inexistent sheet "{sheet_name}"'
            )
        else:
            workbook[sheet_name].title = new_sheet_name
            _workbook_content = list(map(
                lambda v: v.title,
                workbook.worksheets
            ))
            workbook.save(filepath)
            print(
                f'\t> [ Sheet Rename ] Sheet "{
                    sheet_name}"" renamed as "{new_sheet_name}"'
            )
            return workbook
    except:
        print(
            f'\t> [ Sheet Rename ] ❌ Error on rename inexistent sheet "{
                sheet_name}" as "{new_sheet_name}"'
        )


def delete_spreadsheet_sheets(filepath, sheet_names):
    '''delete_spreadsheet_sheets Deletes multiple sheets from a Spreadsheet

    Args:
                                    filepath (str): path of the excel file
                                    sheet_names (list): list of the spreadsheet sheet names

    Returns:
                                    Workbook: Modified Workbook
    '''
    print('\n[[ OpenPyXl ] Automate excel spreadsheet - Delete Spreadsheet sheets ]')
    workbook = openpyxl.load_workbook(filepath)
    print('bforedelete ? workbook:', workbook.worksheets)

    for sheet_name in sheet_names:
        try:
            print(f'\t> [ Sheet Deletion ] Deleting "{sheet_name}"')
            del workbook[sheet_name]

        except:
            print('\t\t> Could not delete sheet "{sheet_name}"')

    workbook.save(filepath)
    print('delete ? workbook:', workbook.worksheets)
    return workbook


# -------------------------------- CELLS FOCUS ------------------------------- #

def manipulate_cells(filepath, sheet_name, values):
    '''manipulate_cells Manipulate cells

    Args:
                    filepath (str): filepath to the local excel
                    values (list): list of values representing the cells
    '''
    print(
        '\n\n[[ OpenPyXl ] Automate excel spreadsheet - Manipulate Spreadsheet Sheet Cells ]')

    # workbook = spreadsheet
    print('\n\t-->path', filepath, '\n')
    workbook = openpyxl.load_workbook(filepath)
    print('\t--> requested excel sheet:', sheet_name)
    print('\t--> current sheets:', workbook.sheetnames)
    try:
        if not (filepath.exists() and sheet_name in workbook.sheetnames):
            print('Sheet ${sheet_name} does not exist, creating')
            workbook.create_sheet(sheet_name)

        worksheet = workbook[sheet_name]
        # ------------------------------ CELL BY ADDRESS ----------------------------- #
        if not worksheet['a1'].value:
            worksheet['a1'] = 'Hello world'

        # spreadsheet sheet cell
        cell_A1_details = worksheet['a1']

        cell_A1_value = cell_A1_details.value

        # reverse string
        reversed_str = cell_A1_value[::-1]

        cell_A1_details = worksheet['a1']

        worksheet['a2'] = reversed_str

        # --------------------------- CELL BY COORDINATION --------------------------- #

        worksheet.cell(3, 1, reversed_str + '_duplicated')

        # Save the workbook
        workbook.save(filepath)

    except Exception as error:
        print(error)


# ---------------------------------------------------------------------------- #
#                                  CELLS RANGE                                 #
# ---------------------------------------------------------------------------- #
def get_and_adjust_cells_range(filepath, sheet_name, range_stx):
    '''get_and_adjust_cells_range Get's a spreadsheet range from a given excel file

    Args:
                                    filepath (Path): local file address
                                    sheet_name (str): sheet name to get cell from
                                    range_stx (str): Range Syntax - Ex: 'A1:L5'
    '''
    print(
        '\n\n[[ OpenPyXl ] Automate excel spreadsheet - Get Spreadsheet Sheet Cells Range ]')
    # ----------------------------- READ CELLS RANGE ----------------------------- #
    try:
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook[sheet_name]
        cells_range = worksheet[range_stx]
        # return cells_range
    except:
        print('Error getting range', range_stx)

    # ------------------------ WRITE / UPDATE CELLS RANGE ------------------------ #
    ''' Cell range structure: tuple [ row items ] of tuples [ cell items ]'''

    # modify path to "generated" path
    filepath_str = str(filepath)
    generated_filepath_str = filepath_str.replace('samples', 'generated')
    generated_path = Path(generated_filepath_str)

    # go through cells and adjust
    count_str = 0
    count_int = 0
    for row in cells_range:
        for cell in row:
            cell_value_type = type(cell.value)
            if cell_value_type is not int:
                cell.value = ''
                count_str += 1
            elif cell_value_type is int:
                if cell.value < 0:
                    cell.value = 0
                    count_int += 1
                elif cell.value > 10:
                    cell.value = 10
                    count_int += 1
                else:
                    pass

    # TODO: save workbook into new path
    print(f'\t> Modified:\n\t|---> {count_str} cell with string\n\t|---> {
        count_int} cell with integer\n')
    workbook.save(generated_path)
    return cells_range


def do_excel_from_data(filepath, data, sheet_name=None,):
    '''do_excel_from_data From a list of list data, create a sheet.
        Generated file - local : generated/data_to_excel.xlsx
        - from data builds sheet file - determine n rows and n columns

    Args:
            filepath (Path): path to store to
            sheet_name (str): name of the spreadsheet sheet
            data (list): list of list e.g.: [[],[],[]]
    '''

    # workbook creation or retrieval
    workbook, worksheet, filename = create_local_excel_file(
        filepath, 'Products'
    )
    worksheet = workbook.active if not sheet_name else workbook[sheet_name]

    # -------------------------- DATA TO CELLS STRUCTURE ------------------------- #
    # --------------------------- Not scalable version --------------------------- #
    first_file_row_values = [
        excelCell.value for excelCell in list(worksheet.rows)[0]
    ]
    are_same_headers = first_file_row_values == data[0]

    # Generates once the file and write data to excel sheet
    if not are_same_headers:
        for (data_row_idx, data_row) in enumerate(data):
            for (col_idx, col) in enumerate(data_row):
                value = data[data_row_idx][col_idx]
                workbook.save(filepath)
        print('\tData written successfully.')
    else:
        print('\t> Nothing was written: data to write already exist.')

    print('\t> See File:', filepath)
