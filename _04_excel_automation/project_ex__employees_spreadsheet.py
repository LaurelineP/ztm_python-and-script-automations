''' Project example Employees Spreadsheet
		Generate a spreadsheet representing employees
		with each sheet representing one employees details
'''
# ------------------------------ PRIVATE IMPORTS ----------------------------- #
import re as _re
from pathlib import Path as _Path

import openpyxl as _openpyxl
from inner_module_utils import load_custom_utils as _load_custom_utils

_log_header, _log, _log_object = _load_custom_utils()


# ---------------------------------------------------------------------------- #
#                                     LOGIC                                    #
# ---------------------------------------------------------------------------- #
# --------------------- CREATE ONE EMPLOYEE SPREADSHEET ---------------------- #
def _create_one_employees_spreadsheet(excel_path: _Path):
    '''create_employees_spreadsheet Create a new employees spreadsheet
            [ from scratch ] - not reusing existing logic personally added ( practice )
    Args:
            excel_path(_type_): excel file path
            content(list, optional): excel sheet content list. Defaults to [].
    '''
    _log('EMPLOYEES SPREADSHEET - CREATE')
    # Creates one file employees spreadsheet
    if not _Path(excel_path).exists():
        new_spreadsheet = _openpyxl.Workbook()
        new_spreadsheet.save(excel_path)
        print(
            f'\t > [ CREATE EMPLOYEE SPREADSHEET ] \n\t "{
                excel_path.name}" file will be created.'
        )
        return new_spreadsheet
    else:
        print(f'\t > [ CREATE EMPLOYEE SPREADSHEET ] \n\t ▶️ "{
            excel_path.name}" file already exists.'
        )
        return _openpyxl.load_workbook(excel_path)


# --------------------------- ADD SHEETS & CONTENT --------------------------- #
def _add_sheet_content(filepath, employee_name, contents):
    '''add_sheet_content Add content to a sheet

    Args:
            filepath (str): filepath name to add sheet to
            employee_name (str): employee name used as a sheet name
            contents (list): list of cells position / value
    '''
    _log(f'Sheet Creation "{employee_name}"')

    # Sheet loading import file contents
    workbook = _openpyxl.load_workbook(filepath)

    # Sheet creation if needed
    worksheets_titles = [ws.title for ws in workbook.worksheets]
    if employee_name not in worksheets_titles:
        workbook.create_sheet(employee_name)
        workbook.save(filepath)

    # Check content and add content to sheet
    has_contents = bool(len(contents))

    cell_pattern = r'^[a-z]\d+$'
    has_valid_contents = all(
        # - should contains a list of 2
        # - and its first item should be of a cell pattern
        len(c) == 2 and _re.search(cell_pattern, c[0])
        for c in contents
    )

    # Writes to the employee sheet the provided content
    if has_contents and has_valid_contents:
        for cell, value in contents:
            # Overrides only if value is not the same
            workbook[employee_name][cell] = value

            workbook.save(filepath)
            print(f'\t> Successfully added cell: \n\t |_______ {
                cell}: {value}\n')

    else:
        print(
            f'Please provide items under the format [ <Cell-Address>, <value> ] {
                contents}'
        )
    return workbook


def generate_employees(filepath, employees, employees_cells_collection):
    '''generate_employees   Generate the whole employees spreadsheets and content

    Args:
            filepath (Path): spreadsheet path
            employees (List[str]): employees list
            employees_cells_collection (List[List[List[str]]]): collection of all
            the employees cells
            - a list to contain them all
            - inner list (level 1) - corresponding to an employee cells collection
            - inner - inner tuple (level 2) - corresponding to a cell (<cell-address>, <value>)

    Returns:
            _type_: SpreadSheet
    '''
    _log('Generate Employees')
    # Creates file
    spreadsheet = _create_one_employees_spreadsheet(filepath)

    # Create a sheet with content per employees
    for employee, employee_cells_list in zip(employees, employees_cells_collection):
        # Create a sheet per employee
        _add_sheet_content(filepath, employee, employee_cells_list)

    print('Spreadsheets sheets:', spreadsheet.sheetnames)

    return spreadsheet
